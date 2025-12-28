import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta, time
import io
import plotly.graph_objects as go
import socket
import platform
import requests
import uuid

# --- نظام مراقبة العمليات (المخفي) ---
if 'admin_logs' not in st.session_state:
    st.session_state.admin_logs = []

# توليد Session ID وهمي للزائر عشان يحس بالخطر
if 'session_id' not in st.session_state:
    st.session_state.session_id = str(uuid.uuid4())[:8].upper()

# إعداد الصفحة
st.set_page_config(page_title="Youssef Pasha - The Ultimate Fix", layout="wide")

# تصميم الواجهة بالألوان المتوهجة (Neon Style)
st.markdown("""
    <style>
    .stApp { background-color: #0e1117; }
    .main-title {
        text-align: center; color: #00f2ff; font-size: 3rem; font-weight: bold;
        text-shadow: 0 0 10px #00f2ff, 0 0 20px #00f2ff; padding: 20px;
    }
    .stFileUploader {
        border: 2px solid #00f2ff; border-radius: 15px; padding: 20px;
        background-color: #161b22; box-shadow: 0 0 15px rgba(0, 242, 255, 0.2);
    }
    .stButton>button {
        width: 100%; background: linear-gradient(45deg, #00f2ff, #7000ff);
        color: white; border: none; border-radius: 10px; padding: 15px;
        font-size: 1.2rem; font-weight: bold; transition: 0.3s;
    }
    .stButton>button:hover { transform: translateY(-2px); box-shadow: 0 0 20px #00f2ff; }
    
    /* ستايل الشريط الجانبي للمبرمج */
    [data-testid="stSidebar"] { background-color: #161b22; border-right: 2px solid #7000ff; }

    /* ستايل كارت بيانات المستخدم */
    .user-card {
        border: 1px solid #7000ff;
        border-radius: 10px;
        padding: 15px;
        background-color: rgba(22, 27, 34, 0.8);
        box-shadow: 0 0 10px rgba(112, 0, 255, 0.3);
        margin-bottom: 20px;
    }
    .user-info-text {
        font-family: 'Courier New', Courier, monospace;
        color: #00f2ff;
        font-size: 14px;
        margin-bottom: 5px;
    }

    .ticker-wrapper {
        position: fixed; bottom: 0; left: 0; width: 100%;
        background: rgba(112, 0, 255, 0.8); color: #00f2ff; padding: 10px 0;
        z-index: 999; border-top: 2px solid #00f2ff; overflow: hidden;
    }
    .ticker-text {
        display: inline-block; white-space: nowrap; font-weight: bold;
        animation: ticker 15s linear infinite;
    }
    @keyframes ticker { 0% { transform: translateX(100%); } 100% { transform: translateX(-100%); } }
    </style>
    
    <div class="ticker-wrapper">
        <div class="ticker-text">
             🚀 تم تطوير النظام بواسطة المبرمج يوسف باشا | Youssef Pasha Fixer V2.0 | جاري تحليل بيانات الزوار.. النظام آمن 🚀
        </div>
    </div>
    """, unsafe_allow_html=True)

# --- دالة جلب بيانات الزائر (Hacker Style) ---
def get_visitor_info():
    try:
        # محاولة جلب الـ IP الخارجي
        ip = requests.get('https://api64.ipify.org?format=json').json()['ip']
    except:
        ip = "Hidden Proxy"
    
    system_info = platform.system() + " " + platform.release()
    node_name = socket.gethostname()
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    return ip, system_info, node_name, current_time

# --- لوحة التحكم السرية للمبرمج يوسف ---
with st.sidebar:
    st.markdown("<h1 style='color: #00f2ff;'>🛡️ الإدارة</h1>", unsafe_allow_html=True)
    password = st.text_input("باسورد المبرمج يوسف", type="password")
    
    if password == "youssef123":
        st.success("أهلاً بك يا مبرمج يوسف")
        st.markdown("### 📊 راصد العمليات")
        if st.session_state.admin_logs:
            for log in reversed(st.session_state.admin_logs):
                st.info(log)
        else:
            st.write("لا توجد عمليات مسجلة حالياً.")
        
        if st.button("مسح السجل"):
            st.session_state.admin_logs = []
            st.rerun()

st.markdown('<p class="main-title">🚀 نظام الإكسيل - يوسف باشا</p>', unsafe_allow_html=True)

# --- تقسيم الشاشة: بيانات الزائر + الرفع ---
main_col1, main_col2 = st.columns([1, 3]) # عمود صغير للبيانات وعمود كبير للرفع

with main_col1:
    # عرض بيانات الزائر بشكل تقني
    ip, sys_info, node, time_now = get_visitor_info()
    st.markdown(f"""
    <div class="user-card">
        <h4 style="color: #7000ff; border-bottom: 1px solid #7000ff;">👁️ بيانات اتصالك</h4>
        <p class="user-info-text">📡 <b>IP:</b> {ip}</p>
        <p class="user-info-text">💻 <b>OS:</b> {sys_info}</p>
        <p class="user-info-text">🆔 <b>Session:</b> {st.session_state.session_id}</p>
        <p class="user-info-text">⏰ <b>Time:</b> {time_now}</p>
        <p class="user-info-text" style="color: #00ff00;">✅ Connection Secure</p>
    </div>
    """, unsafe_allow_html=True)

with main_col2:
    uploaded_file = st.file_uploader("ارفع ملف الإكسيل هنا", type=["xlsx"])

def force_to_time(val):
    if isinstance(val, (time, datetime)):
        return val if isinstance(val, time) else val.time()
    if isinstance(val, str):
        val = val.strip()
        for fmt in ["%H:%M", "%I:%M %p", "%H:%M:%S"]:
            try: return datetime.strptime(val, fmt).time()
            except: continue
    return None

if uploaded_file is not None:
    # تسجيل عملية الرفع
    log_entry = f"📂 IP: {ip} | رفع ملف: {uploaded_file.name} | {datetime.now().strftime('%H:%M:%S')}"
    if log_entry not in st.session_state.admin_logs:
        st.session_state.admin_logs.append(log_entry)

    st.info("📦 تم استلام الملف بنجاح، جاهز للاختراق!")
    
    if st.button("تعديل كافة السجلات المفتوحة (9 ساعات)"):
        with st.spinner('جاري فحص البيانات وتعديل الساعات...'):
            wb = openpyxl.load_workbook(uploaded_file, data_only=True)
            sheet = wb.active
            wb_write = openpyxl.load_workbook(uploaded_file)
            sheet_write = wb_write.active
            
            updates = 0
            total_rows = sheet.max_row - 1
            no_fill = PatternFill(fill_type=None)

            for r in range(2, sheet.max_row + 1):
                c_val = sheet.cell(row=r, column=3).value
                d_val = sheet.cell(row=r, column=4).value
                t_in = force_to_time(c_val)
                
                if t_in and (d_val is None or str(d_val).strip() in ["", "0", "00:00:00", "None"]):
                    dt_in = datetime.combine(datetime.today(), t_in)
                    dt_out = dt_in + timedelta(hours=9)
                    sheet_write.cell(row=r, column=4).value = dt_out.strftime('%H:%M')
                    sheet_write.cell(row=r, column=7).value = 9.0
                    sheet_write.cell(row=r, column=4).fill = no_fill
                    sheet_write.cell(row=r, column=3).fill = no_fill
                    updates += 1

            output = io.BytesIO()
            wb_write.save(output)
            
            if updates > 0:
                # تسجيل النتيجة
                st.session_state.admin_logs.append(f"✅ IP: {ip} | عدل {updates} سطر")
                
                col1, col2 = st.columns([1, 2])
                with col1:
                    st.success(f"✅ تم تعديل {updates} سطر!")
                    st.download_button(label="📥 تحميل الملف المعدل", data=output.getvalue(), file_name="Youssef_Final_9Hours.xlsx")

                with col2:
                    fig = go.Figure(go.Indicator(
                        mode = "gauge+number", value = updates,
                        title = {'text': "عدد السجلات المعدلة", 'font': {'color': "#00f2ff", 'size': 24}},
                        gauge = {'axis': {'range': [None, max(total_rows, updates + 10)], 'tickcolor': "#00f2ff"},
                                 'bar': {'color': "#7000ff"}, 'bordercolor': "#00f2ff"}
                    ))
                    fig.update_layout(paper_bgcolor='rgba(0,0,0,0)', font={'color': "#00f2ff"})
                    st.plotly_chart(fig, use_container_width=True)
                st.balloons()
            else:
                st.error("البرنامج مش شايف خانات ناقصة.")
